from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout as auth_logout
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Case, DecimalField, F, Q, Sum, Value, When
from django.db.models.functions import Coalesce
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.decorators.clickjacking import xframe_options_sameorigin
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from urllib.parse import urlencode
import re

from .forms import ConstructionStatusReportForm, DocumentForm, PlanningBudgetForm, PlanningProjectForm
from .models import ConstructionStatusReport, Document, DocumentScan, PlanningBudget, PlanningProject


def _normalize_excel_header(header_value):
    text = str(header_value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _combine_header_text(primary_header, secondary_header):
    primary = _normalize_excel_header(primary_header)
    secondary = _normalize_excel_header(secondary_header)
    if primary and secondary:
        return f"{primary} {secondary}"
    return primary or secondary


def _to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


def _to_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("%", "")
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    return None


def _to_integer(value):
    decimal_value = _to_decimal(value)
    if decimal_value is None:
        return None
    try:
        return int(decimal_value)
    except (ValueError, OverflowError):
        return None


def _to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _read_excel_construction_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return 0, 0, "Excel import requires `openpyxl`."

    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except InvalidFileException:
        return 0, 0, "Unsupported Excel format. Please upload a valid .xlsx file."
    except Exception:
        return 0, 0, "Unable to read the uploaded Excel file."

    worksheet = workbook.active
    all_rows = list(worksheet.iter_rows(values_only=True))
    if not all_rows:
        return 0, 0, "The Excel file is empty."

    header_map = {
        "project name": "project_name",
        "project title": "project_name",
        "project name title": "project_name",
        "project": "project_name",
        "location": "location",
        "mun": "mun",
        "municipality": "mun",
        "contractor": "contractor",
        "contract cost": "contract_cost",
        "ntp date": "ntp_date",
        "cd": "cd",
        "c d": "cd",
        "contract period cd": "cd",
        "contract period c d": "cd",
        "original expiry date": "original_expiry_date",
        "contract period original expiry date": "original_expiry_date",
        "additional cd": "additional_cd",
        "addl cd": "additional_cd",
        "add l c d": "additional_cd",
        "contract period additional cd": "additional_cd",
        "contract period addl cd": "additional_cd",
        "contract period add l c d": "additional_cd",
        "revised expiry date": "revised_expiry_date",
        "contract period revised expiry date": "revised_expiry_date",
        "date completed": "date_completed",
        "revised contract cost": "revised_contract_cost",
        "previous": "status_previous",
        "status previous": "status_previous",
        "current": "status_current",
        "status current": "status_current",
        "status december 2025 previous": "status_previous",
        "status december 2025 current": "status_current",
        "percent of time elapsed": "percent_time_elapsed",
        "of time elapsed": "percent_time_elapsed",
        "time elapsed": "percent_time_elapsed",
        "slippage": "slippage_percent",
        "slippage percent": "slippage_percent",
        "slippage percent ": "slippage_percent",
        "remarks": "remarks",
    }

    def nearest_secondary_value(start_row_index, column_index, lookahead=3):
        for offset in range(1, lookahead + 1):
            check_row_index = start_row_index + offset
            if check_row_index >= len(all_rows):
                break
            check_row = all_rows[check_row_index] or ()
            if column_index >= len(check_row):
                continue
            value = check_row[column_index]
            if _normalize_excel_header(value):
                return value
        return ""

    header_row_index = None
    column_indexes = {}
    best_score = -1
    scan_limit = min(len(all_rows), 30)
    core_fields = {
        "project_name",
        "location",
        "contractor",
        "contract_cost",
        "ntp_date",
        "cd",
        "original_expiry_date",
        "additional_cd",
        "status_previous",
        "status_current",
        "percent_time_elapsed",
        "slippage_percent",
    }

    for row_index in range(scan_limit):
        candidate_indexes = {}
        row_values = all_rows[row_index] or ()
        if not any(_normalize_excel_header(value) for value in row_values):
            continue
        for col_index, header in enumerate(row_values):
            secondary = nearest_secondary_value(row_index, col_index, lookahead=3)
            normalized_header = _normalize_excel_header(header)
            combined_header = _combine_header_text(header, secondary)
            field_name = header_map.get(combined_header) or header_map.get(normalized_header)
            if field_name and field_name not in candidate_indexes:
                candidate_indexes[field_name] = col_index
        if "project_name" not in candidate_indexes:
            continue
        score = sum(1 for field in core_fields if field in candidate_indexes)
        if score > best_score:
            best_score = score
            header_row_index = row_index
            column_indexes = candidate_indexes

    if header_row_index is None:
        return 0, 0, "Missing required `Project Name` column in Excel."

    created_count = 0
    skipped_count = 0

    meta_project_labels = {
        "project name",
        "prepared by",
        "checked by",
        "noted by",
        "reviewed by",
        "certified correct",
        "approved by",
        "billing",
        "percentage of billing",
        "status of billing on process payment received",
    }

    for row in all_rows[header_row_index + 1:]:
        if not row:
            continue

        project_col = column_indexes.get("project_name")
        project_name = _to_text(row[project_col]) if project_col is not None and project_col < len(row) else ""
        normalized_project = _normalize_excel_header(project_name)
        if (
            (not project_name)
            or (normalized_project in meta_project_labels)
            or normalized_project.startswith("prepared by")
            or normalized_project.startswith("billing")
            or normalized_project.startswith("percentage of billing")
            or normalized_project.startswith("status of billing")
        ):
            skipped_count += 1
            continue

        def cell(field):
            col_index = column_indexes.get(field)
            if col_index is None or col_index >= len(row):
                return None
            return row[col_index]

        ConstructionStatusReport.objects.create(
            project_name=project_name,
            location=_to_text(cell("location")),
            mun=_to_text(cell("mun")),
            contractor=_to_text(cell("contractor")),
            contract_cost=_to_decimal(cell("contract_cost")),
            ntp_date=_to_date(cell("ntp_date")),
            cd=_to_integer(cell("cd")),
            original_expiry_date=_to_date(cell("original_expiry_date")),
            additional_cd=_to_integer(cell("additional_cd")),
            revised_expiry_date=_to_text(cell("revised_expiry_date")),
            date_completed=_to_date(cell("date_completed")),
            revised_contract_cost=_to_decimal(cell("revised_contract_cost")),
            status_previous=_to_text(cell("status_previous")),
            status_current=_to_text(cell("status_current")),
            percent_time_elapsed=_to_decimal(cell("percent_time_elapsed")),
            slippage_percent=_to_decimal(cell("slippage_percent")),
            remarks=_to_text(cell("remarks")),
        )
        created_count += 1

    return created_count, skipped_count, ""


def logout_view(request):
    if request.method == "POST":
        auth_logout(request)
        return render(
            request,
            "registration/logout.html",
            {"login_url": reverse("login")},
        )
    return redirect("login")


@login_required
def home(request):
    item_type_choices = [
        ("document", "Document"),
        ("task", "Task"),
        ("project", "Project (go to Projects)"),
    ]
    division_choices = Document.DIVISION_CHOICES
    selected_division = Document.DIV_CONSTRUCTION
    selected_item_type = "document"
    title = ""
    description = ""
    show_new_item_modal = False
    new_item_error = ""

    if request.method == "POST" and request.POST.get("action") == "create_home_item":
        selected_item_type = request.POST.get("item_type", "document").strip().lower()
        selected_division = request.POST.get("division", "").strip()
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        show_new_item_modal = True

        valid_item_types = {value for value, _ in item_type_choices}
        valid_divisions = {value for value, _ in division_choices}

        if selected_item_type not in valid_item_types:
            new_item_error = "Invalid item type."
        elif not title:
            new_item_error = "Title is required."
        elif selected_item_type == "document":
            if selected_division not in valid_divisions:
                new_item_error = "Invalid division selected."
            elif not _table_exists(Document):
                new_item_error = "Document table is not ready yet. Run migrations first."
            else:
                Document.objects.create(
                    document_name=title,
                    doc_type=Document.TYPE_OTHER,
                    division=selected_division,
                    status=Document.STATUS_DRAFT,
                    description=description,
                    created_by=request.user,
                )
                return redirect("admin_div_dashboard")
        elif selected_item_type == "task":
            return redirect("maintinance_task_management")
        elif selected_item_type == "project":
            return redirect(f"{reverse('planning_div_dashboard')}?tab=ppa")

        if selected_item_type != "document" and selected_division not in valid_divisions:
            new_item_error = "Invalid division selected."

    total_documents = 0
    completed_documents = 0
    ongoing_documents = 0
    total_cost_value = Decimal("0")
    recent_projects = []

    if _table_exists(Document):
        documents_qs = Document.objects.all()
        total_documents = documents_qs.count()
        completed_documents = documents_qs.filter(
            status__in=[Document.STATUS_APPROVED, Document.STATUS_CLOSED]
        ).count()
        ongoing_documents = documents_qs.exclude(
            status__in=[Document.STATUS_APPROVED, Document.STATUS_CLOSED]
        ).count()
        total_cost_value = documents_qs.aggregate(
            total=Coalesce(
                Sum(
                    Case(
                        When(revised_contract_amount__isnull=False, then=F("revised_contract_amount")),
                        default=Coalesce(
                            F("contract_amount"),
                            Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                            output_field=DecimalField(max_digits=14, decimal_places=2),
                        ),
                        output_field=DecimalField(max_digits=14, decimal_places=2),
                    )
                ),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        )["total"] or Decimal("0")

    if _table_exists(PlanningProject):
        recent_projects = list(PlanningProject.objects.order_by("-id")[:6])

    def _format_currency_short(amount):
        amount = Decimal(amount or 0)
        absolute = abs(amount)
        if absolute >= Decimal("1000000000"):
            return f"PHP {amount / Decimal('1000000000'):.1f}B"
        if absolute >= Decimal("1000000"):
            return f"PHP {amount / Decimal('1000000'):.1f}M"
        return f"PHP {amount:,.0f}"

    context = {
        "item_type_choices": item_type_choices,
        "division_choices": division_choices,
        "new_item_selected_type": selected_item_type,
        "new_item_selected_division": selected_division,
        "new_item_title": title,
        "new_item_description": description,
        "show_new_item_modal": show_new_item_modal,
        "new_item_error": new_item_error,
        "total_documents": total_documents,
        "completed_documents": completed_documents,
        "ongoing_documents": ongoing_documents,
        "total_cost_display": _format_currency_short(total_cost_value),
        "recent_projects": recent_projects,
    }
    return render(request, "home.html", context)


@login_required
@xframe_options_sameorigin
def planning_div_dashboard(request):
    active_tab = request.GET.get("tab", "budget")
    if active_tab not in {"budget", "ppa", "timeline"}:
        active_tab = "budget"

    selected_fund = request.GET.get("fund", "").strip()
    if not _table_exists(PlanningBudget) or not _table_exists(PlanningProject):
        context = _planning_fallback_context(active_tab=active_tab, selected_fund=selected_fund)
        return render(request, "Planning Division/planning_Div.html", context)

    if not PlanningBudget.objects.exists():
        PlanningBudget.objects.create(
            name="20% Development Fund FY 2026",
            fund=PlanningBudget.FUND_20_DEV,
            fiscal_year="FY 2026",
            total_budget=625105288,
            allocated_amount=0,
            status=PlanningBudget.STATUS_APPROVED,
        )
        PlanningBudget.objects.create(
            name="SEF FY 2026",
            fund=PlanningBudget.FUND_SEF,
            fiscal_year="FY 2026",
            total_budget=62994000,
            allocated_amount=0,
            status=PlanningBudget.STATUS_APPROVED,
        )

    budgets = list(PlanningBudget.objects.all())
    projects = PlanningProject.objects.all()
    show_budget_modal = False
    editing_budget = None
    show_project_modal = False
    project_form = PlanningProjectForm(initial={"fund": selected_fund or PlanningBudget.FUND_20_DEV})

    edit_budget_id = request.GET.get("edit_budget")
    if edit_budget_id:
        editing_budget = PlanningBudget.objects.filter(id=edit_budget_id).first()
        if editing_budget:
            show_budget_modal = True

    ppa_search = request.GET.get("q", "").strip()
    ppa_status = request.GET.get("status", "").strip()
    ppa_fund = request.GET.get("fund_filter", selected_fund).strip()

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "delete_budget":
            delete_id = request.POST.get("delete_id")
            if delete_id:
                PlanningBudget.objects.filter(id=delete_id).delete()
            return redirect(f"{request.path}?tab=budget")

        if action == "create_budget":
            budget_form = PlanningBudgetForm(request.POST)
            show_budget_modal = True
            editing_budget = None
            if budget_form.is_valid():
                budget_form.save()
                return redirect(f"{request.path}?tab=budget")
        elif action == "update_budget":
            budget_id = request.POST.get("budget_id")
            editing_budget = PlanningBudget.objects.filter(id=budget_id).first()
            budget_form = PlanningBudgetForm(request.POST, instance=editing_budget)
            show_budget_modal = True
            if budget_form.is_valid():
                budget_form.save()
                return redirect(f"{request.path}?tab=budget")
        elif action == "create_project":
            project_form = PlanningProjectForm(request.POST)
            show_project_modal = True
            if project_form.is_valid():
                project = project_form.save()
                return redirect(
                    f"{request.path}?tab=ppa&fund={project.fund}"
                )
            budget_form = PlanningBudgetForm(instance=editing_budget) if editing_budget else PlanningBudgetForm()
        elif action == "delete_project":
            project_id = request.POST.get("project_id")
            if project_id:
                PlanningProject.objects.filter(id=project_id).delete()
            return redirect(f"{request.path}?tab=ppa&fund={selected_fund or ''}")
        else:
            budget_form = PlanningBudgetForm(instance=editing_budget) if editing_budget else PlanningBudgetForm()
    else:
        budget_form = PlanningBudgetForm(instance=editing_budget) if editing_budget else PlanningBudgetForm()

    project_totals = {key: Decimal("0") for key, _ in PlanningBudget.FUND_CHOICES}
    for project in projects:
        project_totals[project.fund] = project_totals.get(project.fund, Decimal("0")) + (project.budget_amount or 0)

    for budget in budgets:
        budget.computed_allocated = project_totals.get(budget.fund, Decimal("0"))

    total_allocated = sum((b.allocated_value for b in budgets), Decimal("0"))
    total_remaining = sum((b.remaining_amount for b in budgets), Decimal("0"))

    ppa_projects = projects
    if ppa_search:
        ppa_projects = ppa_projects.filter(project_title__icontains=ppa_search)
    if ppa_status:
        ppa_projects = ppa_projects.filter(status=ppa_status)
    if ppa_fund:
        ppa_projects = ppa_projects.filter(fund=ppa_fund)

    ppa_total = projects.count()
    ppa_approved = projects.filter(status=PlanningProject.STATUS_APPROVED).count()
    ppa_for_review = projects.filter(status=PlanningProject.STATUS_FOR_REVIEW).count()
    ppa_total_cost = sum((p.budget_amount for p in projects), Decimal("0"))

    context = {
        "active_tab": active_tab,
        "selected_fund": selected_fund,
        "budgets": budgets,
        "budget_form": budget_form,
        "show_budget_modal": show_budget_modal,
        "editing_budget": editing_budget,
        "project_form": project_form,
        "show_project_modal": show_project_modal,
        "ppa_projects": ppa_projects,
        "ppa_search": ppa_search,
        "ppa_status": ppa_status,
        "ppa_fund": ppa_fund,
        "ppa_total": ppa_total,
        "ppa_approved": ppa_approved,
        "ppa_for_review": ppa_for_review,
        "ppa_total_cost": ppa_total_cost,
        "project_status_choices": PlanningProject.STATUS_CHOICES,
        "fund_choices": PlanningBudget.FUND_CHOICES,
        "total_budgets": len(budgets),
        "total_allocated": total_allocated,
        "total_remaining": total_remaining,
    }
    return render(request, "Planning Division/planning_Div.html", context)


@login_required
@xframe_options_sameorigin
def admin_div_dashboard(request):
    active_tab = request.GET.get("tab", "documents")
    open_create_modal = request.GET.get("new_document") == "1"
    if active_tab not in {"documents", "billing"}:
        active_tab = "documents"
    if not _table_exists(Document):
        context = _admin_fallback_context(active_tab=active_tab)
        return render(request, "Admin/admin_div.html", context)

    form = DocumentForm()
    show_modal = False
    editing_document = None

    edit_document_id = request.GET.get("edit_document")
    if active_tab == "documents" and edit_document_id:
        editing_document = Document.objects.filter(id=edit_document_id).first()
        if editing_document:
            form = DocumentForm(instance=editing_document)
            show_modal = True
    elif active_tab == "documents" and open_create_modal:
        show_modal = True

    if request.method == "POST":
        action = request.POST.get("action", "create").strip()

        if action == "update_status":
            document_id = request.POST.get("document_id")
            new_status = request.POST.get("status", "").strip()
            valid_statuses = {value for value, _ in Document.STATUS_CHOICES}
            if document_id and new_status in valid_statuses:
                document = Document.objects.select_related("project").filter(id=document_id).first()
                if document:
                    document.status = new_status
                    is_billing_record = bool(document.billing_type) or document.doc_type == Document.TYPE_BILLING_PACKET

                    if (
                        is_billing_record
                        and new_status == Document.STATUS_APPROVED
                        and _table_exists(PlanningProject)
                        and _table_exists(PlanningBudget)
                    ):
                        mapped_fund = _map_billing_type_to_fund(document.billing_type)
                        if mapped_fund:
                            allocation_amount = (
                                document.revised_contract_amount
                                or document.contract_amount
                                or Decimal("0")
                            )

                            planning_project = document.project
                            if planning_project:
                                planning_project.fund = mapped_fund
                                planning_project.budget_amount = allocation_amount
                                planning_project.status = PlanningProject.STATUS_APPROVED
                                planning_project.save(update_fields=["fund", "budget_amount", "status"])
                            else:
                                planning_project = PlanningProject.objects.filter(
                                    project_title=document.document_name,
                                    fund=mapped_fund,
                                ).order_by("-id").first()
                                if planning_project:
                                    planning_project.budget_amount = allocation_amount
                                    planning_project.status = PlanningProject.STATUS_APPROVED
                                    planning_project.save(update_fields=["budget_amount", "status"])
                                else:
                                    planning_project = PlanningProject.objects.create(
                                        project_title=document.document_name,
                                        fund=mapped_fund,
                                        budget_amount=allocation_amount,
                                        status=PlanningProject.STATUS_APPROVED,
                                    )

                            document.project = planning_project
                            document.division = Document.DIV_PLANNING

                    document.save(update_fields=["status", "project", "division"])

            next_url = request.POST.get("next", "").strip()
            if next_url:
                return redirect(next_url)
            return redirect(f"{reverse('admin_div_dashboard')}?tab={active_tab}")

        if active_tab == "documents":
            if action == "delete_document":
                delete_id = request.POST.get("delete_id")
                if delete_id:
                    Document.objects.filter(id=delete_id).delete()
                return redirect("admin_div_dashboard")

            if action == "delete_scan":
                document_id = request.POST.get("document_id")
                scan_id = request.POST.get("scan_id")
                if document_id and scan_id:
                    scan = DocumentScan.objects.filter(id=scan_id, document_id=document_id).first()
                    if scan:
                        if scan.file:
                            scan.file.delete(save=False)
                        scan.delete()
                if document_id:
                    return redirect(f"{reverse('admin_div_dashboard')}?tab=documents&edit_document={document_id}")
                return redirect("admin_div_dashboard")

            if action == "replace_scan":
                document_id = request.POST.get("document_id")
                scan_id = request.POST.get("scan_id")
                replacement_file = request.FILES.get("replacement_scan_file")
                if document_id and scan_id and replacement_file:
                    scan = DocumentScan.objects.select_related("document").filter(id=scan_id, document_id=document_id).first()
                    if scan:
                        if scan.file:
                            scan.file.delete(save=False)
                        scan.file = replacement_file
                        scan.project = scan.document.project
                        scan.uploaded_by = request.user
                        scan.save(update_fields=["file", "project", "uploaded_by"])
                if document_id:
                    return redirect(f"{reverse('admin_div_dashboard')}?tab=documents&edit_document={document_id}")
                return redirect("admin_div_dashboard")

            if action in {"create", "update"}:
                instance = None
                form_payload = request.POST
                if action == "update":
                    document_id = request.POST.get("document_id")
                    if document_id:
                        instance = Document.objects.filter(id=document_id).first()
                        editing_document = instance
                    if instance:
                        # Keep the payload mutable so we can only backfill missing fields.
                        form_payload = request.POST.copy()
                        if "project" not in form_payload:
                            form_payload["project"] = str(instance.project_id) if instance.project_id else ""
                        if not form_payload.get("division"):
                            form_payload["division"] = instance.division
                        if not form_payload.get("status"):
                            form_payload["status"] = instance.status
                        if not form_payload.get("doc_type"):
                            form_payload["doc_type"] = instance.doc_type

                form = DocumentForm(form_payload, instance=instance)
                show_modal = True
                if form.is_valid():
                    target_division = form.cleaned_data.get("division")
                    uploaded_scans = [file for file in request.FILES.getlist("scanned_files") if file]
                    has_existing_scans = bool(instance and instance.scans.exists())
                    requires_quality_scan = target_division == Document.DIV_QUALITY

                    if requires_quality_scan and not uploaded_scans and not has_existing_scans:
                        form.add_error(
                            None,
                            "For Quality Division submissions, upload at least one scanned request letter.",
                        )
                    else:
                        document = form.save(commit=False)
                        if instance is None:
                            document.created_by = request.user
                        document.save()
                        for uploaded_file in uploaded_scans:
                            DocumentScan.objects.create(
                                document=document,
                                project=document.project,
                                file=uploaded_file,
                                uploaded_by=request.user,
                            )
                        if document.division == Document.DIV_QUALITY:
                            return redirect("quality_div_dashboard")
                        return redirect("admin_div_dashboard")

    search = request.GET.get("q", "").strip()
    division = request.GET.get("division", "").strip()
    status = request.GET.get("status", "").strip()

    documents_qs = Document.objects.all()

    if search:
        search_l = search.lower()
        matching_types = [
            value
            for value, label in Document.TYPE_CHOICES
            if search_l in label.lower() or search_l in value.replace("_", " ").lower()
        ]
        matching_divisions = [
            value
            for value, label in Document.DIVISION_CHOICES
            if search_l in label.lower() or search_l in value.replace("_", " ").lower()
        ]
        matching_statuses = [
            value
            for value, label in Document.STATUS_CHOICES
            if search_l in label.lower() or search_l in value.replace("_", " ").lower()
        ]
        documents_qs = documents_qs.filter(
            Q(slip_ref_no__icontains=search)
            | Q(document_name__icontains=search)
            | Q(billing_type__icontains=search)
            | Q(description__icontains=search)
            | Q(doc_type__in=matching_types)
            | Q(division__in=matching_divisions)
            | Q(status__in=matching_statuses)
        )

    documents_filtered = documents_qs
    if division:
        documents_filtered = documents_filtered.filter(division=division)
    if status:
        documents_filtered = documents_filtered.filter(status=status)

    # Billing tab mirrors document records so any status update in
    # Document Register is immediately reflected in Billing.
    billing_qs = documents_qs
    if status:
        billing_qs = billing_qs.filter(status=status)

    if active_tab == "billing":
        paginator = Paginator(billing_qs, 8)
    else:
        paginator = Paginator(documents_filtered, 8)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Show contractor in admin document register for contract-type documents.
    # Contractor data is sourced from construction reports by matching project title.
    reports_map = {}
    if _table_exists(ConstructionStatusReport):
        project_names = {
            ((doc.project.project_title if doc.project else doc.document_name) or "").strip()
            for doc in page_obj.object_list
            if doc.doc_type == Document.TYPE_CONTRACT
        }
        project_names = {name for name in project_names if name}
        if project_names:
            reports = ConstructionStatusReport.objects.filter(project_name__in=project_names).values(
                "project_name", "contractor"
            )
            for report in reports:
                key = (report.get("project_name") or "").strip().lower()
                contractor = (report.get("contractor") or "").strip()
                if key and contractor:
                    reports_map[key] = contractor

    for doc in page_obj.object_list:
        contractor_name = (doc.contractor_name or "").strip()
        if contractor_name:
            doc.contractor_display = contractor_name
            continue

        doc.contractor_display = "-"
        if doc.doc_type == Document.TYPE_CONTRACT:
            lookup_title = ((doc.project.project_title if doc.project else doc.document_name) or "").strip().lower()
            if lookup_title:
                doc.contractor_display = reports_map.get(lookup_title, "-")

    billing_base = Document.objects.all()

    context = {
        "active_tab": active_tab,
        "form": form,
        "show_modal": show_modal,
        "editing_document": editing_document,
        "documents": page_obj.object_list if active_tab == "documents" else documents_filtered[:0],
        "billing_records": page_obj.object_list if active_tab == "billing" else billing_qs[:0],
        "page_obj": page_obj,
        "search": search,
        "selected_division": division,
        "selected_status": status,
        "division_choices": Document.DIVISION_CHOICES,
        "status_choices": Document.STATUS_CHOICES,
        "total_documents": Document.objects.count(),
        "for_review_count": Document.objects.filter(status=Document.STATUS_FOR_REVIEW).count(),
        "processing_count": Document.objects.filter(status=Document.STATUS_PROCESSING).count(),
        "open_issues_count": Document.objects.filter(status=Document.STATUS_OPEN).count(),
        "total_billing_records": billing_base.count(),
        "billing_for_review_count": billing_base.filter(status=Document.STATUS_FOR_REVIEW).count(),
        "billing_processing_count": billing_base.filter(status=Document.STATUS_PROCESSING).count(),
        "billing_approved_count": billing_base.filter(status=Document.STATUS_APPROVED).count(),
    }
    return render(request, "Admin/admin_div.html", context)


@login_required
@xframe_options_sameorigin
def maintinance_div_dashboard(request):
    return _render_maintinance_dashboard(request)


@login_required
@xframe_options_sameorigin
def construction_div_dashboard(request):
    embedded_mode = bool(request.GET.get("embedded") or request.POST.get("embedded"))
    page_param = (request.POST.get("page") or request.GET.get("page") or "").strip()
    redirect_params = {}
    if embedded_mode:
        redirect_params["embedded"] = "1"
    if page_param:
        redirect_params["page"] = page_param
    redirect_url = reverse("construction_div_dashboard")
    if redirect_params:
        redirect_url = f"{redirect_url}?{urlencode(redirect_params)}"
    import_success = request.GET.get("imported", "").strip()
    import_skipped = request.GET.get("skipped", "").strip()
    import_error = request.GET.get("import_error", "").strip()

    if not _table_exists(ConstructionStatusReport):
        context = {
            "reports": [],
            "report_form": ConstructionStatusReportForm(),
            "show_report_modal": False,
            "editing_report": None,
            "db_not_ready": True,
        }
        return render(request, "Construction Division/construction_div.html", context)

    reports_qs = ConstructionStatusReport.objects.all()
    reports_paginator = Paginator(reports_qs, 10)
    reports_page = reports_paginator.get_page(page_param or 1)
    reports = reports_page.object_list
    show_report_modal = False
    editing_report = None

    edit_id = request.GET.get("edit")
    if edit_id:
        editing_report = ConstructionStatusReport.objects.filter(id=edit_id).first()
        if editing_report:
            show_report_modal = True

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "delete":
            delete_id = request.POST.get("delete_id")
            if delete_id:
                ConstructionStatusReport.objects.filter(id=delete_id).delete()
            return redirect(redirect_url)

        if action == "bulk_delete":
            selected_ids = request.POST.get("selected_ids", "")
            id_list = []
            for value in selected_ids.split(","):
                value = value.strip()
                if not value:
                    continue
                try:
                    id_list.append(int(value))
                except ValueError:
                    continue
            if id_list:
                ConstructionStatusReport.objects.filter(id__in=id_list).delete()
            return redirect(redirect_url)

        if action == "import_excel":
            uploaded_excel = request.FILES.get("excel_file")
            if not uploaded_excel:
                query = urlencode({"import_error": "Please choose an Excel file."})
            else:
                created_count, skipped_count, error_message = _read_excel_construction_rows(uploaded_excel)
                if error_message:
                    query = urlencode({"import_error": error_message})
                else:
                    query = urlencode({"imported": str(created_count), "skipped": str(skipped_count)})
            separator = "&" if "?" in redirect_url else "?"
            return redirect(f"{redirect_url}{separator}{query}")

        if action in {"create", "update"}:
            instance = None
            update_error = None
            if action == "update":
                report_id = request.POST.get("report_id")
                if report_id:
                    instance = ConstructionStatusReport.objects.filter(id=report_id).first()
                    editing_report = instance
                if instance is None:
                    update_error = "The selected report no longer exists."
            report_form = ConstructionStatusReportForm(request.POST, instance=instance)
            show_report_modal = True
            if update_error:
                report_form.add_error(None, update_error)
            elif report_form.is_valid():
                report_form.save()
                return redirect(redirect_url)
        else:
            report_form = ConstructionStatusReportForm()
    else:
        report_form = (
            ConstructionStatusReportForm(instance=editing_report)
            if editing_report
            else ConstructionStatusReportForm()
        )

    context = {
        "reports": reports,
        "page_obj": reports_page,
        "report_form": report_form,
        "show_report_modal": show_report_modal,
        "editing_report": editing_report,
        "import_success": import_success,
        "import_skipped": import_skipped,
        "import_error": import_error,
    }
    return render(request, "Construction Division/construction_div.html", context)


@login_required
@xframe_options_sameorigin
def quality_div_dashboard(request):
    admin_projects = []
    admin_incoming_documents = []
    if _table_exists(Document):
        seen_titles = set()
        incoming_qs = (
            Document.objects.filter(division=Document.DIV_QUALITY)
            .select_related("project")
            .prefetch_related("scans")
            .order_by("created_at", "id")
        )

        for doc in incoming_qs:
            project_title = ""
            if doc.project_id and doc.project:
                project_title = (doc.project.project_title or "").strip()
            if not project_title:
                project_title = (doc.document_name or "").strip()
            if project_title:
                title_key = project_title.lower()
                if title_key not in seen_titles:
                    seen_titles.add(title_key)
                    admin_projects.append(project_title)

            doc_date = doc.date_received_from_admin or doc.date_released_to_admin
            received_date = doc.date_received_from_admin or doc.created_at.date()
            latest_scan = doc.scans.first()
            request_letter_url = ""
            request_letter_name = ""
            if latest_scan and latest_scan.file:
                request_letter_url = latest_scan.file.url
                request_letter_name = latest_scan.file.name.split("/")[-1]
            admin_incoming_documents.append(
                {
                    "id": doc.id,
                    "received_from": "Admin Division",
                    "doc_letter_date": doc_date,
                    "particulars": doc.get_doc_type_display(),
                    "details": doc.description or "-",
                    "document_no": doc.slip_ref_no or f"DOC-{doc.id:04d}",
                    "project_name": project_title or "-",
                    "location": (doc.location or "").strip() or "-",
                    "request_letter_url": request_letter_url,
                    "request_letter_name": request_letter_name,
                    "route": "Incoming",
                    "released_by": "-",
                    "date_released": None,
                    "received_by": "Quality Division",
                    "date_received": received_date,
                    "transmitted_to": doc.get_division_display(),
                    "remarks": "-",
                    "initial_status": doc.get_status_display(),
                    "owner_representative": "-",
                    "contact": "-",
                }
            )

    context = {
        "admin_projects": admin_projects,
        "admin_incoming_documents": admin_incoming_documents,
    }
    return render(request, "Quality Division/quality_div.html", context)


@login_required
@xframe_options_sameorigin
def my_assignments(request):
    return render(request, "My Assignments/my_Assigments.html")


@login_required
@xframe_options_sameorigin
def history_page(request):
    return render(request, "History/history.html")


@login_required
@xframe_options_sameorigin
def projects_dashboard(request):
    query = request.GET.get("q", "").strip()
    division_choices = [
        (Document.DIV_ADMIN, "Admin Division"),
        (Document.DIV_PLANNING, "Planning Division"),
        (Document.DIV_CONSTRUCTION, "Construction Division"),
        (Document.DIV_QUALITY, "Quality Division"),
        (Document.DIV_MAINTENANCE, "Maintenance Division"),
    ]
    valid_divisions = {value for value, _ in division_choices}
    selected_division = request.GET.get("division", Document.DIV_ADMIN).strip().lower()
    if selected_division not in valid_divisions:
        selected_division = Document.DIV_ADMIN

    counts_by_division = {value: 0 for value, _ in division_choices}
    locations_by_division = {value: set() for value, _ in division_choices}
    selected_projects = []
    project_folders = []

    if _table_exists(Document):
        def _canonical_project_key(raw_value):
            # Canonical key for de-duplicating project titles across divisions.
            # It normalizes case, punctuation/symbols, and repeated whitespace.
            return _normalize_excel_header(raw_value)

        def _quality_entry_title_and_key(document):
            # Keep Projects grouping consistent with Quality dashboard project derivation.
            title = ""
            if document.project_id and document.project:
                title = (document.project.project_title or "").strip()
            if not title:
                title = (document.document_name or "").strip()
            if not title:
                title = f"Document #{document.id}"
            compact_title = re.sub(r"\s+", " ", title).strip()
            return compact_title, _canonical_project_key(compact_title)

        def _entry_title_and_key(document):
            title = ""
            if document.project_id and document.project:
                title = (document.project.project_title or "").strip()
            if not title:
                title = (document.document_name or f"Document #{document.id}").strip()
            normalized_title = re.sub(r"\s+", " ", title).strip()
            return normalized_title, _canonical_project_key(normalized_title)

        division_entry_keys = {value: set() for value, _ in division_choices}
        for div_key, _ in division_choices:
            division_docs = Document.objects.filter(division=div_key).select_related("project")
            for doc in division_docs:
                if div_key == Document.DIV_QUALITY:
                    _, entry_key = _quality_entry_title_and_key(doc)
                else:
                    _, entry_key = _entry_title_and_key(doc)
                division_entry_keys[div_key].add(entry_key)
                location_value = (doc.location or "").strip()
                if location_value:
                    locations_by_division[div_key].add(location_value.lower())
            counts_by_division[div_key] = len(division_entry_keys[div_key])

        construction_reports_qs = ConstructionStatusReport.objects.none()
        if _table_exists(ConstructionStatusReport):
            construction_reports_qs = ConstructionStatusReport.objects.exclude(project_name__isnull=True).exclude(project_name__exact="")
            for report in construction_reports_qs.only("project_name", "location"):
                project_key = _canonical_project_key(report.project_name)
                if project_key:
                    division_entry_keys[Document.DIV_CONSTRUCTION].add(project_key)
                location_value = (report.location or "").strip()
                if location_value:
                    locations_by_division[Document.DIV_CONSTRUCTION].add(location_value.lower())
            counts_by_division[Document.DIV_CONSTRUCTION] = len(division_entry_keys[Document.DIV_CONSTRUCTION])

    location_meta_by_division = {}
    for div_key, _ in division_choices:
        location_count = len(locations_by_division.get(div_key, set()))
        if location_count:
            label = "location" if location_count == 1 else "locations"
            location_meta_by_division[div_key] = f"{location_count} {label}"
        else:
            location_meta_by_division[div_key] = "No location yet"

        documents_qs = (
            Document.objects.filter(division=selected_division)
            .select_related("project")
            .prefetch_related("scans")
            .order_by("-created_at")
        )
        if query:
            documents_qs = documents_qs.filter(
                Q(document_name__icontains=query)
                | Q(description__icontains=query)
                | Q(project__project_title__icontains=query)
            )

        folders_map = {}
        for doc in documents_qs:
            if selected_division == Document.DIV_QUALITY:
                entry_title, entry_key = _quality_entry_title_and_key(doc)
            else:
                entry_title, entry_key = _entry_title_and_key(doc)
            bucket = folders_map.setdefault(
                entry_key,
                {
                    "project_title": entry_title,
                    "document_count": 0,
                    "files": [],
                    "documents": [],
                },
            )
            bucket["document_count"] += 1
            bucket["documents"].append(
                {
                    "name": doc.document_name,
                    "status": doc.get_status_display(),
                    "division": doc.get_division_display(),
                    "location": (doc.location or "").strip(),
                    "created_at": doc.created_at,
                }
            )
            for scan in doc.scans.all():
                bucket["files"].append(scan)

        if selected_division == Document.DIV_CONSTRUCTION and _table_exists(ConstructionStatusReport):
            if query:
                construction_reports_qs = construction_reports_qs.filter(
                    Q(project_name__icontains=query)
                    | Q(location__icontains=query)
                    | Q(contractor__icontains=query)
                    | Q(remarks__icontains=query)
                )
            for report in construction_reports_qs.order_by("-created_at"):
                project_title = (report.project_name or "").strip()
                if not project_title:
                    continue
                entry_key = _canonical_project_key(project_title)
                bucket = folders_map.setdefault(
                    entry_key,
                    {
                        "project_title": project_title,
                        "document_count": 0,
                        "files": [],
                        "documents": [],
                    },
                )
                bucket["document_count"] += 1
                status_segments = []
                if report.status_previous:
                    status_segments.append(f"Prev: {report.status_previous}")
                if report.status_current:
                    status_segments.append(f"Curr: {report.status_current}")
                if report.slippage_percent is not None:
                    status_segments.append(f"Slippage: {report.slippage_percent}%")
                bucket["documents"].append(
                    {
                        "name": "Construction Status Report",
                        "status": " | ".join(status_segments) if status_segments else "Recorded",
                        "division": "Construction Division",
                        "location": (report.location or "").strip(),
                        "created_at": report.created_at,
                    }
                )

        consolidated_folders = {}
        for _, bucket in folders_map.items():
            canonical_title_key = _canonical_project_key(bucket.get("project_title", ""))
            merged_bucket = consolidated_folders.setdefault(
                canonical_title_key,
                {
                    "project_title": bucket.get("project_title", ""),
                    "document_count": 0,
                    "files": [],
                    "documents": [],
                },
            )
            merged_bucket["document_count"] += bucket.get("document_count", 0)
            merged_bucket["files"].extend(bucket.get("files", []))
            merged_bucket["documents"].extend(bucket.get("documents", []))
            if len(bucket.get("project_title", "")) > len(merged_bucket.get("project_title", "")):
                merged_bucket["project_title"] = bucket.get("project_title", "")

        for _, bucket in sorted(consolidated_folders.items(), key=lambda item: item[1]["project_title"].lower()):
            seen_document_rows = set()
            deduped_documents = []
            for doc_row in bucket["documents"]:
                row_key = (
                    (doc_row.get("name") or "").strip().casefold(),
                    (doc_row.get("status") or "").strip().casefold(),
                    (doc_row.get("division") or "").strip().casefold(),
                    str(doc_row.get("created_at") or ""),
                    (doc_row.get("location") or "").strip().casefold(),
                )
                if row_key in seen_document_rows:
                    continue
                seen_document_rows.add(row_key)
                deduped_documents.append(doc_row)
            bucket["documents"] = deduped_documents
            bucket["document_count"] = len(deduped_documents)

            selected_projects.append(
                {
                    "title": bucket["project_title"],
                    "meta": f"{bucket['document_count']} document(s) | {len(bucket['files'])} scan(s)",
                }
            )
            project_folders.append(bucket)

    context = {
        "query": query,
        "selected_division": selected_division,
        "selected_division_label": dict(division_choices).get(selected_division, "Division"),
        "counts_by_division": counts_by_division,
        "location_meta_by_division": location_meta_by_division,
        "selected_projects": selected_projects,
        "project_folders": project_folders,
    }
    return render(request, "Projects/projects.html", context)


@login_required
@xframe_options_sameorigin
def maintinance_task_management(request):
    return _render_maintinance_tasks(request)


@login_required
@xframe_options_sameorigin
def maintinance_contractor_management(request):
    return _render_maintinance_contractors(request)


@login_required
@xframe_options_sameorigin
def maintenance_div_dashboard(request):
    return _render_maintinance_dashboard(request)


@login_required
@xframe_options_sameorigin
def maintenance_task_management(request):
    return _render_maintinance_tasks(request)


@login_required
@xframe_options_sameorigin
def maintenance_contractor_management(request):
    return _render_maintinance_contractors(request)


def _render_maintinance_dashboard(request):
    return render(request, "Maintinance Division/maintinance_management.html")


def _render_maintinance_tasks(request):
    return render(request, "Maintinance Division/task_management.html")


def _render_maintinance_contractors(request):
    return render(request, "Maintinance Division/contractor_management.html")


def _table_exists(model):
    table_name = model._meta.db_table
    try:
        return table_name in connection.introspection.table_names()
    except Exception:
        return False


def _planning_fallback_context(active_tab="budget", selected_fund=""):
    return {
        "active_tab": active_tab,
        "selected_fund": selected_fund,
        "budgets": [],
        "budget_form": PlanningBudgetForm(),
        "show_budget_modal": False,
        "editing_budget": None,
        "project_form": PlanningProjectForm(initial={"fund": selected_fund or PlanningBudget.FUND_20_DEV}),
        "show_project_modal": False,
        "ppa_projects": [],
        "ppa_search": "",
        "ppa_status": "",
        "ppa_fund": "",
        "ppa_total": 0,
        "ppa_approved": 0,
        "ppa_for_review": 0,
        "ppa_total_cost": Decimal("0"),
        "project_status_choices": PlanningProject.STATUS_CHOICES,
        "fund_choices": PlanningBudget.FUND_CHOICES,
        "total_budgets": 0,
        "total_allocated": Decimal("0"),
        "total_remaining": Decimal("0"),
        "db_not_ready": True,
    }


def _map_billing_type_to_fund(billing_type):
    normalized = (billing_type or "").strip().lower()
    if "20" in normalized and "development" in normalized:
        return PlanningBudget.FUND_20_DEV
    if "sef" in normalized:
        return PlanningBudget.FUND_SEF
    return ""


def _admin_fallback_context(active_tab="documents"):
    empty_page = Paginator([], 8).get_page(1)
    return {
        "active_tab": active_tab,
        "form": DocumentForm(),
        "show_modal": False,
        "editing_document": None,
        "documents": [],
        "billing_records": [],
        "page_obj": empty_page,
        "search": "",
        "selected_division": "",
        "selected_status": "",
        "division_choices": Document.DIVISION_CHOICES,
        "status_choices": Document.STATUS_CHOICES,
        "total_documents": 0,
        "for_review_count": 0,
        "processing_count": 0,
        "open_issues_count": 0,
        "total_billing_records": 0,
        "billing_for_review_count": 0,
        "billing_processing_count": 0,
        "billing_approved_count": 0,
        "db_not_ready": True,
    }
